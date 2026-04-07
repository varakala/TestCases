from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


HEADERS = [
    "Test Case ID",
    "Endpoint",
    "Tags",
    "Test Scenario",
    "Test Description",
    "Preconditions",
    "Test Steps",
    "Test Data",
    "Expected Result",
    "Priority",
    "Test Type",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
CELL_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

PRIORITY_COLORS = {
    "High": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "Medium": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "Low": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
}

COLUMN_WIDTHS = [15, 30, 20, 30, 45, 25, 40, 30, 40, 10, 15]


def export_to_excel(all_test_cases, output_path):
    """Export test cases to a formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Write headers
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    # Set column widths
    for col, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else "A" + chr(64 + col - 26)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:K1"

    # Write test cases
    row = 2
    tc_counter = 1
    for test_case in all_test_cases:
        global_id = f"TC_{tc_counter:04d}"
        values = [
            global_id,
            test_case.get("endpoint", ""),
            test_case.get("tags", ""),
            test_case.get("test_scenario", ""),
            test_case.get("test_description", ""),
            test_case.get("preconditions", ""),
            test_case.get("test_steps", ""),
            test_case.get("test_data", ""),
            test_case.get("expected_result", ""),
            test_case.get("priority", ""),
            test_case.get("test_type", ""),
        ]

        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=str(value) if value else "")
            cell.font = CELL_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

        # Color-code priority
        priority = test_case.get("priority", "")
        if priority in PRIORITY_COLORS:
            priority_cell = ws.cell(row=row, column=10)
            priority_cell.fill = PRIORITY_COLORS[priority]

        row += 1
        tc_counter += 1

    # Add summary sheet
    _add_summary_sheet(wb, all_test_cases)

    wb.save(output_path)
    return output_path


def _add_summary_sheet(wb, all_test_cases):
    """Add a summary sheet with counts by endpoint and priority."""
    ws = wb.create_sheet("Summary")

    ws.cell(row=1, column=1, value="Test Case Summary").font = Font(bold=True, size=14)
    ws.cell(row=3, column=1, value="Total Test Cases:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=len(all_test_cases))

    # Priority breakdown
    ws.cell(row=5, column=1, value="By Priority").font = Font(bold=True, size=12)
    priorities = {}
    for tc in all_test_cases:
        p = tc.get("priority", "Unknown")
        priorities[p] = priorities.get(p, 0) + 1

    row = 6
    for priority, count in sorted(priorities.items()):
        ws.cell(row=row, column=1, value=priority)
        ws.cell(row=row, column=2, value=count)
        row += 1

    # By endpoint
    row += 1
    ws.cell(row=row, column=1, value="By Endpoint").font = Font(bold=True, size=12)
    row += 1
    endpoints = {}
    for tc in all_test_cases:
        ep = tc.get("endpoint", "Unknown")
        endpoints[ep] = endpoints.get(ep, 0) + 1

    for ep, count in sorted(endpoints.items()):
        ws.cell(row=row, column=1, value=ep)
        ws.cell(row=row, column=2, value=count)
        row += 1

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
